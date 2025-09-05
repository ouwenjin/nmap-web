from __future__ import annotations
import sys
import os
import re
import textwrap
import argparse
import unicodedata
import logging
import xml.etree.ElementTree as ET
from typing import List, Dict, Any, Optional

# 第三方库（请确保已安装：pandas, openpyxl, tqdm）
try:
    import pandas as pd
except Exception as e:
    print("缺少 pandas，请通过 `pip install pandas` 安装。")
    raise

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except Exception as e:
    print("缺少 openpyxl，请通过 `pip install openpyxl` 安装。")
    raise

try:
    from tqdm import tqdm
except Exception:
    # 如果没有 tqdm，提供替代实现（非常简陋）
    def tqdm(iterable, **kwargs):
        return iterable

# ------------------------- 常量 & 日志 -------------------------
AUTHOR = "zhkali"
REPOS = [
    "https://github.com/ouwenjin/nmap-web",
    "https://gitee.com/zhkali/nmap-web",
]

ANSI = {
    "reset": "\033[0m",
    "bold": "\033[1m",
    "cyan": "\033[36m",
    "green": "\033[32m",
    "yellow": "\033[33m",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("merge.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger("security_toolkit")

# ------------------------- 辅助：宽字符与颜色处理 -------------------------
_ansi_re = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')

def supports_color() -> bool:
    """检测终端是否支持 ANSI 颜色（Windows 做基本兼容判断）"""
    if sys.platform.startswith("win"):
        return os.getenv("ANSICON") is not None or "WT_SESSION" in os.environ or sys.stdout.isatty()
    return sys.stdout.isatty()

_COLOR = supports_color()

def strip_ansi(s: str) -> str:
    return _ansi_re.sub("", s)

def visible_width(s: str) -> int:
    """计算字符串在终端中的可见宽度（中文宽度计为 2）"""
    s2 = strip_ansi(s)
    w = 0
    for ch in s2:
        if unicodedata.combining(ch):
            continue
        ea = unicodedata.east_asian_width(ch)
        w += 2 if ea in ("F", "W") else 1
    return w

def pad_visible(s: str, target_visible_len: int) -> str:
    cur = visible_width(s)
    if cur >= target_visible_len:
        return s
    return s + " " * (target_visible_len - cur)

# ------------------------- Banner（来自 1.py 的改进版） -------------------------
def _banner_lines() -> List[str]:
    big_name = r"""
   ███████╗██╗  ██╗██╗  ██╗ █████╗ ██╗      ██╗        
   ╚══███╔╝██║  ██║██║ ██╔╝██╔══██╗██║      ██║        
     ███╔╝ ███████║█████╔╝ ███████║██║      ██║        
    ███╔╝  ██╔══██║██╔═██╗ ██╔══██║██║      ██║        
   ███████╗██║  ██║██║  ██╗██║  ██║███████╗ ██║       
   ╚══════╝╚═╝  ╚═╝╚═╝  ╚═╝╚═╝  ╚═╝╚══════╝ ╚═╝        
"""
    art = textwrap.dedent(big_name)
    art_lines = [ln.rstrip("\n") for ln in art.splitlines() if ln.strip() != ""]
    return art_lines + [""] + [f"作者： {AUTHOR}"] + REPOS

def print_banner(use_unicode: bool = True, outer_margin: int = 0, inner_pad: int = 1):
    """
    打印带边框的作者横幅，保留对中文/宽字符的对齐支持。
    参数：
      - use_unicode: 是否使用 Unicode 盒绘字符（默认 True）
      - outer_margin: 左侧空格数
      - inner_pad: 内容与竖线的左右内边距
    """
    if use_unicode:
        tl, tr, bl, br, hor, ver = "┌", "┐", "└", "┘", "─", "│"
    else:
        tl, tr, bl, br, hor, ver = "+", "+", "+", "+", "-", "|"

    c_reset = ANSI.get("reset", "")
    c_bold = ANSI.get("bold", "")
    c_cyan = ANSI.get("cyan", "")
    c_green = ANSI.get("green", "")
    c_yellow = ANSI.get("yellow", "")

    raw_lines = _banner_lines()
    colored = []
    for ln in raw_lines:
        if ln.startswith("作者"):
            colored.append((c_bold + c_green + ln + c_reset) if _COLOR else ln)
        elif ln.startswith("http"):
            colored.append((c_yellow + ln + c_reset) if _COLOR else ln)
        else:
            colored.append((c_bold + c_cyan + ln + c_reset) if _COLOR else ln)

    content_max = max((visible_width(x) for x in colored), default=0)
    padded_lines = [pad_visible(ln, content_max) for ln in colored]
    total_inner = inner_pad * 2 + content_max
    width = total_inner + 2  # 两侧竖线
    top = tl + (hor * (width - 2)) + tr
    bottom = bl + (hor * (width - 2)) + br
    pad = " " * max(0, outer_margin)

    # 顶部
    if _COLOR and use_unicode:
        print(pad + (c_cyan + top + c_reset))
    else:
        print(pad + top)

    # 内容行
    left_bar = (c_cyan + ver + c_reset) if _COLOR else ver
    right_bar = left_bar
    for pl in padded_lines:
        line_content = (" " * inner_pad) + pl + (" " * inner_pad)
        print(pad + left_bar + line_content + right_bar)

    # 底部
    if _COLOR and use_unicode:
        print(pad + (c_cyan + bottom + c_reset))
    else:
        print(pad + bottom)

# ------------------------- Excel 提取（来自 web.py，改进） -------------------------
def extract_from_xlsx_interactive(input_file: Optional[str] = None, output_file: str = "web探测.txt"):
    """
    从给定（或由用户选择）的 Excel 中提取 IP:端口 并去重写入 txt。
    如果 input_file 为 None，则列出当前目录的 .xlsx 供用户选择。
    """
    # 如果没有传入文件名，则列出选择
    if not input_file:
        excel_files = [f for f in os.listdir(".") if f.lower().endswith(".xlsx")]
        if not excel_files:
            print("当前目录下没有找到任何 .xlsx 文件。")
            return
        print("\n可用的 Excel 文件：")
        for i, fn in enumerate(excel_files, 1):
            print(f"  {i}) {fn}")
        choice = input(f"请选择要解析的文件（1-{len(excel_files)}，默认 1）：").strip()
        choice = choice if choice else "1"
        try:
            idx = int(choice) - 1
            if idx < 0 or idx >= len(excel_files):
                raise ValueError("选择超出范围")
            input_file = excel_files[idx]
        except Exception as e:
            print("无效选择，已取消。")
            return

    # 读取 Excel 并提取
    results = set()
    try:
        wb = openpyxl.load_workbook(input_file, data_only=True)
    except Exception as e:
        print(f"打开 Excel 文件失败：{input_file}，错误：{e}")
        return

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        # 假设第一列为 IP，第二列为 端口/协议（从第 2 行开始）
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            ip = str(row[0]).strip() if row[0] is not None else ""
            port_proto = str(row[1]).strip() if row[1] is not None else ""
            if not ip or not port_proto:
                continue
            port = port_proto.split("/")[0]
            # 简单判断 IPv6（含 : 且没有 .）则用 [ip]:port 格式
            if ":" in ip and not ip.count("."):
                address = f"[{ip}]:{port}"
            else:
                address = f"{ip}:{port}"
            results.add(address)

    # 输出
    try:
        with open(output_file, "w", encoding="utf-8") as f:
            for addr in sorted(results):
                f.write(addr + "\n")
        print(f"提取完成，去重后共 {len(results)} 条，结果已保存到 {output_file}")
    except Exception as e:
        print(f"写入文件失败：{e}")

# ------------------------- Nmap XML -> Excel（来自 nmap-xml-xlsx.py） -------------------------
# 危险端口/服务集（用于标记）
dangerous_ports = {
    20,21,23,25,53,69,111,110,2049,143,137,135,139,389,445,161,
    512,513,514,873,1433,1521,1529,3306,3389,5000,5432,
    5900,5901,5902,6379,7001,888,9200,9300,11211,27017,27018
}
dangerous_services = {
    'ftp','telnet','smtp','dns','smb','snmp','rsync','oracle','mysql','mysqlx',
    'mariadb','rdp','postgresql','vnc','redis','weblogic_server','elasticsearch',
    'elasticsearch_transport','memcached','mongodb','mongodb_shard_or_secondary',
    'tftp','nfs','pop3','imap','netbios-ns','msrpc','netbios-ssn','ldap',
    'linux rexec','mssql','oracle db','sybase/db2','ilo','any','oracledb',
    'http','linuxrexec','vnc服务'
}

def is_valid_ip(ip: Optional[str]) -> bool:
    if not ip:
        return False
    ipv4_pattern = r"^(25[0-5]|2[0-4]\d|[01]?\d\d?)(\.(25[0-5]|2[0-4]\d|[01]?\d\d?)){3}$"
    ipv6_pattern = r"^([0-9a-fA-F]{0,4}:){2,7}[0-9a-fA-F]{0,4}$"
    return re.match(ipv4_pattern, ip) is not None or re.match(ipv6_pattern, ip) is not None

def merge_all_xml(output_file: str = "out.xml") -> Optional[str]:
    """
    将当前目录下所有 .xml 文件合并为一个 out.xml（将所有 <host> 节点追加到第一个 xml 的根）
    返回合并后的文件名；如果没有 xml，返回 None。
    """
    xml_files = [f for f in os.listdir(".") if f.lower().endswith(".xml")]
    if not xml_files:
        logger.warning("没有找到 XML 文件，跳过合并。")
        return None

    logger.info(f"开始合并 {len(xml_files)} 个 XML 文件 -> {output_file}")
    try:
        main_tree = ET.parse(xml_files[0])
        main_root = main_tree.getroot()
    except Exception as e:
        logger.error(f"解析 {xml_files[0]} 失败: {e}")
        return None

    for xml_file in xml_files[1:]:
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            for host in root.findall("host"):
                main_root.append(host)
        except Exception as e:
            logger.error(f"合并文件 {xml_file} 出错: {e}")

    try:
        main_tree.write(output_file, encoding="utf-8", xml_declaration=True)
        logger.info(f"XML 合并完成，结果保存为 {output_file}")
        return output_file
    except Exception as e:
        logger.error(f"写出合并文件失败: {e}")
        return None

def parse_nmap_xml(xml_file: str) -> List[Dict[str, Any]]:
    """
    解析合并好的 xml 文件，返回每个端口一条记录的列表。
    每条记录字典包含：IP, 端口/协议, 状态, 服务, 端口用途（空）
    """
    results = []
    if not os.path.exists(xml_file):
        logger.warning(f"文件不存在: {xml_file}")
        return results

    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
    except Exception as e:
        logger.error(f"解析 XML 文件失败: {e}")
        return results

    hosts = root.findall("host")
    for h_index, host in enumerate(tqdm(hosts, desc=f"解析Nmap ({xml_file})", unit="host")):
        addr_elem = host.find("address")
        ip = addr_elem.get("addr") if addr_elem is not None else None
        if ip is not None and not is_valid_ip(ip):
            logger.debug(f"[Nmap] Host#{h_index} IP 可能无效: {ip}")

        for port in host.findall(".//port"):
            proto = port.get("protocol")
            portid = port.get("portid")
            state_elem = port.find("state")
            state = state_elem.get("state") if state_elem is not None else ""
            service_elem = port.find("service")
            service = service_elem.get("name") if service_elem is not None else ""
            results.append({
                "IP": ip,
                "端口/协议": f"{portid}/{proto}",
                "状态": state,
                "服务": service,
                "端口用途": ""
            })
    return results

def auto_dedup(df: pd.DataFrame) -> (pd.DataFrame, str):
    """对关键列去重并返回（去重后的 df, 模式描述）"""
    if df.empty:
        return df, "none"
    before = len(df)
    df.drop_duplicates(subset=["IP", "端口/协议", "服务", "状态", "端口用途"], inplace=True)
    after = len(df)
    return df, f"strict ({before-after} 行被删除)"

def mark_dangerous(df: pd.DataFrame) -> pd.DataFrame:
    """根据危险端口/服务集合标注是否需要开放"""
    def check(row):
        try:
            port = int(str(row["端口/协议"]).split("/")[0])
        except Exception:
            port = None
        service = str(row["服务"]).strip().lower()
        if (port in dangerous_ports) or (service in dangerous_services):
            return "危险端口不允许对外开放"
        return ""
    df["是否必要开放"] = df.apply(check, axis=1)
    return df

def format_excel(file_path: str):
    """对导出的 Excel 做简单美化（列宽、字体、红色标记）"""
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        logger.error(f"打开 Excel 美化失败: {e}")
        return
    ws = wb.active
    font = Font(name="宋体", size=12)
    bold_font = Font(name="宋体", size=12, bold=True)
    red_font = Font(name="宋体", size=12, color="FFFF0000")
    column_widths = {"A":36, "B":12, "C":12, "D":18, "E":11, "F":28}
    for col, width in column_widths.items():
        try:
            ws.column_dimensions[col].width = width
        except Exception:
            pass
    for row in ws.iter_rows():
        for cell in row:
            try:
                cell.font = font
                if cell.row == 1:
                    cell.font = bold_font
                if cell.value == "危险端口不允许对外开放":
                    cell.font = red_font
            except Exception:
                pass
    try:
        wb.save(file_path)
    except Exception as e:
        logger.error(f"保存 Excel 美化失败: {e}")

def run_nmap_xml(output_excel: str = "端口调研表.xlsx") -> Optional[str]:
    """
    完整的 Nmap XML -> Excel 流程：
      1. 合并 xml -> out.xml
      2. 解析 out.xml
      3. 去重、标注
      4. 输出到 output_excel 并美化
      5. 删除 out.xml
    返回 output_excel 的路径（成功）或 None（失败）
    """
    merged = merge_all_xml("out.xml")
    if not merged:
        logger.error("未生成合并的 XML，已停止。")
        return None

    all_results = parse_nmap_xml(merged)
    if not all_results:
        logger.error("未解析到任何端口记录。")
        # 尝试删除临时文件
        try:
            os.remove(merged)
        except Exception:
            pass
        return None

    df = pd.DataFrame(all_results)
    df, mode = auto_dedup(df)
    logger.info(f"自动去重模式：{mode}，最终 {len(df)} 行")
    df = mark_dangerous(df)

    try:
        df.to_excel(output_excel, index=False)
        format_excel(output_excel)
        logger.info(f"处理完成，结果保存为 {output_excel}")
    except Exception as e:
        logger.error(f"保存 Excel 失败: {e}")
        return None
    finally:
        # 清理临时文件
        try:
            if os.path.exists(merged):
                os.remove(merged)
                logger.info(f"已删除临时文件：{merged}")
        except Exception:
            pass

    return os.path.abspath(output_excel)

# ------------------------- 主入口：支持命令行与交互菜单 -------------------------
def interactive_menu():
    print("\n=== 安全工具集菜单 ===")
    print("1) 解析 Nmap XML (生成端口调研表.xlsx)")
    print("2) 提取 Excel IP:端口 (从指定或选择的 Excel 中提取，输出 web探测.txt)")
    print("3) 完整流程 (解析 XML -> 提取 IP:端口)")
    choice = input("请选择功能 (1/2/3, q 退出): ").strip()

    if choice == "1":
        print_banner()
        run_nmap_xml()
    elif choice == "2":
        print_banner()
        extract_from_xlsx_interactive()
    elif choice == "3":
        # 完整流程：每个主要步骤都打印作者信息
        print_banner()
        excel_path = run_nmap_xml()
        if excel_path:
            print_banner()
            # 传入生成的 Excel 路径以避免二次选择
            extract_from_xlsx_interactive(input_file=os.path.basename(excel_path))
    elif choice.lower() == "q":
        print("退出。")
    else:
        print("无效选择。")


def main():
    parser = argparse.ArgumentParser(description="安全工具集 (Banner / Excel 提取 / Nmap XML 解析)")
    subparsers = parser.add_subparsers(dest="cmd", help="子命令")

    # banner 子命令（单独打印）
    p_banner = subparsers.add_parser("banner", help="仅打印横幅")
    p_banner.add_argument("--no-unicode", action="store_true", help="禁用 Unicode 边框")
    p_banner.add_argument("--margin", type=int, default=0, help="左侧外边距")
    p_banner.add_argument("--pad", type=int, default=1, help="内部左右内边距")

    # extract 子命令（非交互，可直接传入文件）
    p_extract = subparsers.add_parser("extract", help="从 Excel 提取 IP:端口（可指定 input_file）")
    p_extract.add_argument("-i", "--input", help="输入 Excel 文件路径（默认为交互选择）")
    p_extract.add_argument("-o", "--output", default="web探测.txt", help="输出文件名")

    # nmap 子命令（直接运行解析流程）
    p_nmap = subparsers.add_parser("nmap", help="解析当前目录下所有 Nmap XML -> 生成端口调研表.xlsx")
    p_nmap.add_argument("-o", "--output", default="端口调研表.xlsx", help="输出 Excel 文件名")

    args = parser.parse_args()

    if args.cmd is None:
        # 无命令参数进入交互式菜单
        interactive_menu()
        return

    # 命令行模式处理
    if args.cmd == "banner":
        # 这里把 no-unicode 与 margin/pad 支持传递到 print_banner（需要额外参数实现）
        use_unicode = not getattr(args, "no_unicode", False)
        # outer_margin/pad 若需要使用可扩展 print_banner 参数
        print_banner(use_unicode=use_unicode, outer_margin=getattr(args, "margin", 0), inner_pad=getattr(args, "pad", 1))
    elif args.cmd == "extract":
        if getattr(args, "input", None):
            print_banner()
            extract_from_xlsx_interactive(input_file=args.input, output_file=args.output)
        else:
            print_banner()
            extract_from_xlsx_interactive(input_file=None, output_file=args.output)
    elif args.cmd == "nmap":
        print_banner()
        run_nmap_xml(output_excel=args.output)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
